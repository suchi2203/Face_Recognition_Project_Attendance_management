from openpyxl import Workbook, load_workbook

# to be used by creater only 


# wb=Workbook('attendance.xlsx')
# wb.create_sheet("January")
# wb.create_sheet("February")
# wb.create_sheet("March")
# wb.create_sheet("April")
# wb.create_sheet("May")
# wb.create_sheet("June")
# wb.create_sheet("July")
# wb.create_sheet("August")
# wb.create_sheet("September")
# wb.create_sheet("October")
# wb.create_sheet("November")
# wb.create_sheet("December")
# wb.save('attendance.xlsx')


# wb=load_workbook('attendance.xlsx')
# months=['January','February','March','April','May','June','July','August','September','October','November','December']
# l=["sno","Id","Name","01","02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31"]
# for month in months:
#     ws=wb[month]
#     if month in (['January','March','May','July','August','October','December']):
#         ws.append(l)
#         print("done")        
#     elif month in ('April','June','September','November'):
#         ws.append(l[:32])  
#         print("done")  
#     else:
#         ws.append(l[:31])
#         print("done")

# wb.save('attendance.xlsx')