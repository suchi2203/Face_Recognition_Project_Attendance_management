from flask import Flask,render_template,Response,request,redirect
import cv2
import face_recognition
import numpy as np
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from openpyxl import Workbook, load_workbook


# getting admin credentials(Id and password) for login
with open('admin_credentials.txt','r') as f:
    cred=f.readlines()          # cred=list of admin credentials,
f.close()
cred=cred[0].split(" ")         # cred[0] = admin id, cred[1] = password

app=Flask(__name__)                    


# linking app to database
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data_reg.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


# Database Model
class Att(db.Model):                         
    sno= db.Column(db.Integer, primary_key=True)         # the primary key
    Id = db.Column(db.Integer,nullable=False)            # Id of user can be roll no of students or id of employees 
    name = db.Column(db.String(20),nullable=False)
    email = db.Column(db.String(40),nullable=False)
    face_encoding = db.Column(db.String(3000))           # face encoding stored in string format
    date_registered = db.Column(db.DateTime, default=datetime.utcnow)
    
    def __repr__(self) -> str:
        return f"{self.Id} {self.name} {self.email}"


# function to capture the video from webcam
def generate_frames():
    global camera                           # global camera provides access outside function as well 
    camera=cv2.VideoCapture(0)
    while True:

        # read the camera frame
        success,frame=camera.read()
        frameS=cv2.resize(frame,(0,0),None,0.25,0.25)
        frameS=cv2.cvtColor(frameS,cv2.COLOR_BGR2RGB)    #convt BGR format to RGB format

        face_curr_frame=face_recognition.face_locations(frameS)
        global encode_curr_frame
        #face encodings of face currently captured by camera
        encode_curr_frame=face_recognition.face_encodings(frameS,face_curr_frame) 
        

        if  not success:
            break
        if len(face_curr_frame)==0:
            # if no faces in frame c
            ret,buffer=cv2.imencode('.jpg',frame)
            frame=buffer.tobytes()
            #passing video in form of continuous
            yield(b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

        else:
            for encode_face,faceloc in zip(encode_curr_frame,face_curr_frame):
                y1,x2,y2,x1=faceloc
                y1,x2,y2,x1=y1*4,x2*4,y2*4,x1*4
                # forming rectangle around detected face
                cv2.rectangle(frame,(x1,y1),(x2,y2),(0,255,255),2)

                ret,buffer=cv2.imencode('.jpg',frame)
                frame=buffer.tobytes()
            
                yield(b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')    


# to route to index.html 
@app.route('/',methods=['POST', 'GET'])
def index():
    return render_template('index.html')


# capture video until capture button isn't pressed and go to index.html as soon as pressed
@app.route('/video',methods=['POST', 'GET'])
def video():
    if request.method=='POST':
        camera.release()
        return render_template('index.html')
    else:
        return Response(generate_frames(),mimetype='multipart/x-mixed-replace; boundary=frame')


# to match the face in frame with the registered face
@app.route('/match_att',methods=['POST', 'GET'])
def match_att():
    # when id is entered
    if request.method=="POST":
        id=request.form['Id']
        person=Att.query.filter_by(Id=id).first()
        
        # if no person with that id is registered
        if person==None:
            message=True        # for message to be displayed
            find_face=True
            return render_template('give_att.html',message=message,find_face=find_face)
        else:
            
            if len(encode_curr_frame)<1:
                message=False
                find_face=False         # no face found in frame
                return render_template('give_att.html',message=message,find_face=find_face)
            else:
                
                camera.release()
                str_encoding=person.face_encoding
                # id registered but face not registered
                if str_encoding=="0":
                    already=False
                    name=person.name
                    no_face_reg=True
                    return render_template('att_continue.html',match=False,name=name,already=already,no_face_reg=no_face_reg)
             
                name=person.name
                list_known_face=[float(x) for x in str_encoding.split(" ")]
                match=face_recognition.compare_faces([list_known_face],encode_curr_frame[0],tolerance=0.45)  #matching faces current v/s registered
                
                # if current face matches to registered face
                if match[0]==True:
                    wb=load_workbook('attendance.xlsx')
                    month=datetime.now().strftime("%B")
                    ws=wb[month]
                    sno=person.sno              #sno is the primary key
                    row_=sno+1                  # row is uniquely identified as sno+1
                    a=datetime.now().date()   # today's date
                    column_=int(str(a)[-2]+str(a)[-1])+3        #first date starts from third column

                    # if attendance already marked
                    if ws.cell(row=row_, column=column_).value =="P":
                        wb.save('attendance.xlsx')
                        already=True           # attendance already marked
                        no_face_reg=False
                        return render_template('att_continue.html',match=match[0],name=name,already=already,no_face_reg=no_face_reg)
                    
                    else:
                        # mark the attendance 
                        ws.cell(row=row_, column=column_).value ="P"
                        wb.save('attendance.xlsx')
                        already=False
                        no_face_reg=False
                        return render_template('att_continue.html',match=match[0],name=name,already=already,no_face_reg=no_face_reg)
                
                else:
                    #if the face doesn't match the registered face
                    no_face_reg=False
                    already=False
                    return render_template('att_continue.html',match=False,name=name,no_face_reg=no_face_reg,already=already)

    else:
        # till id is not entered
        message=False
        find_face=True
        return render_template('give_att.html',message=message,find_face=find_face)

# routing to give_att.html
@app.route('/give_attendance')
def give_attendance():
    message=False
    return render_template('give_att.html',message=message)

# routing to admin_action.html
@app.route('/admin_login',methods=['POST', 'GET'])
def admin_action():
    # if id password are entered
    if request.method=='POST':
        adminId=request.form['admin_Id']
        pswd_=request.form['admin_psw']

        if adminId==cred[0] and pswd_==cred[1]:     # admin credentials matched, redirect to new page
            return render_template('del_or_reg.html')
        else:
            message=True                 # admin credentials didn't match, stay at same page and display message 
            return render_template('admin_action.html',message=message)
    else:
        message=False
        return render_template('admin_action.html',message=message)    


# route to new_regis.html
@app.route('/new_registration')
def new_registration():
    return render_template('new_regis.html')


# to proceed registration, provide form for entry with method=POST
@app.route('/reg_proceed',methods=['POST', 'GET'])
def reg_proceed():
    # as soon as the details are entered
    if request.method=='POST':
        global Id_
        Id_ = request.form['Id']
        name_ = request.form['name']
        email_ = request.form['email']
        personz=Att.query.filter_by(Id=Id_).first()  # create a new object of Att class
        if personz==None:
            global new_person
            new_person = Att(Id=Id_, name=name_ ,email=email_, face_encoding="0")   # face encoding="0" till no data is provided 
            db.session.add(new_person)              # adding the details into database
            db.session.commit()
            SNo=new_person.sno
            wb=load_workbook('attendance.xlsx')
            row_=SNo+1
            # entering the details in all sheets
            sheetnames=["January","February","March","April","May","June","July","August","September","October","November","December"]
            for sheet in wb.sheetnames:
                ws=wb[sheet]
                ws.cell(row=row_, column=1).value =SNo
                ws.cell(row=row_, column=2).value =Id_
                ws.cell(row=row_, column=3).value =name_     
                wb.save('attendance.xlsx')
            message=False
            return render_template('reg_proceed.html',message=message) 
        else:
            #if that id already exists
            message=True  
            return render_template('new_regis.html',message=message)

    else:
        message=False
        return render_template('new_regis.html',message= message)


# routing to video registration
@app.route('/video_reg',methods=['POST', 'GET'])
def video_reg():
    # if capture button is clicked
    if request.method=='POST':
        # if no face detected in frame print message of no face detection
        if len(encode_curr_frame)<1:
            camera.release()
            message=True
            return render_template('reg_proceed.html',message=message)

        a= encode_curr_frame[0]
        camera.release()
        person=Att.query.filter_by(Id=Id_).first()
        # converting ndarray of face encoding to string
        str_a=" ".join(str(x) for x in a.tolist())

        # updating default face encoding "0" to str_a
        person.face_encoding=(str_a)
        name_=person.name
        try:
            db.session.commit()
            return render_template('regis_success.html',name=name_)
        except:
            return "Couldn't perform the action"
    # till capture button isn't clicked
    else:
        return Response(generate_frames(),mimetype='multipart/x-mixed-replace; boundary=frame')


# routing to page containing list of all persons and link to view their annual attendance
@app.route('/view_list_att')
def view_list():
    persons = Att.query.order_by(Att.sno).all()
    return render_template('view_list_att.html',persons=persons)


# routing to page containing list of all persons and link delete them from list i.e. deregister them
@app.route('/view_list_del')
def view_list_del():
    persons = Att.query.order_by(Att.sno).all()
    return render_template('view_list_del.html',persons=persons)


# Removing/Deleting user
@app.route('/remove_user/<int:sno>')
def remove_user(sno):
    person= Att.query.get_or_404(sno)
    db.session.delete(person)           #deleting query
    db.session.commit()
    persons = Att.query.order_by(Att.sno).all()
    return render_template('view_list_del.html',persons=persons)

# route to page showing the annual attendance report of user 
@app.route('/show_att/<int:sno>')
def update(sno):
    person= Att.query.get_or_404(sno)
    wb=load_workbook('attendance.xlsx')
    year=["January","February","March","April","May","June","July","August","September","October","November","December"]
    att=[]              # list which will contain 12 sublists each of which will contain monthly attendance 
    row_=sno+1          
    for month in year:
        ws=wb[month]
        m=[]
        for i in range(4,35):
            a=ws.cell(row=row_,column=i).value
            if a==None:
                if month in ['February','April','June','September','November'] and i==34:
                    m.append("X")                           # putting a X in months days lesser than 31
                elif month=='February' and i==33:           
                    m.append("X")
                else:                                        
                    m.append(" ")                   # a blank if No value at cell
            else:
                m.append(a)
        att.append(m)
    name_=ws.cell(row=row_,column=3).value
    return render_template('show_att.html',listt=att,name=name_)    # the attendance passed to html page as list


if __name__=='__main__':
    app.run(debug=True)

